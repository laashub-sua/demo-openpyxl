import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import colors


def gen_excel(excel_file_path, excel_metadata, excel_datalist, font_condition=None, font_style=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'sheet1'
    for index in range(len(excel_metadata)):  # 生成首行
        index_x_name = chr(ord('A') + index)
        index_y_name = str(1)
        index_name = index_x_name + index_y_name
        ws.column_dimensions[index_x_name].width = 20.0
        ws[index_name] = excel_metadata[index]

    for index in range(len(excel_datalist)):  # 生成数据行
        item = excel_datalist[index]
        ws.append(item)
        index_y = index + 2
        for index_x in range(len(item)):  # A -> Z
            if font_condition and font_condition(index_x, index_y, item, excel_datalist):  # 如果满足设置font的条件则设置font
                ws[chr(ord('A') + index_x) + str(index_y)].font = font_style

    wb.save(excel_file_path)


if __name__ == '__main__':
    # 定义数据模型
    test_excel_metadata = ['国家', '首都']
    test_excel_datalist = [
        ['中国', '北京'],
        ['中国', '北京'],
        ['中国', '北京'],
        ['中国', '北京'],
        ['中国', '北京'],
        ['中国', '北京'],
        ['中国', '北京'],
        ['中国', '北京'],
    ]


    def font_condition(index_x, index_y, item, excel_datalist):
        if index_y % 2 == 0:
            return False
        return True


    gen_excel('test.xls', test_excel_metadata, test_excel_datalist, font_condition=font_condition,
              font_style=Font(color=colors.RED))
